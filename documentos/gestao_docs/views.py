from .decorators import check_permission
from django.shortcuts import render, get_object_or_404, redirect
from .models import Funcionario, Documento, LocalMobilizacao, GestorLocal, Perfil
from .forms import LocalMobilizacaoForm, FuncionarioForm, DocumentoForm, GestorLocalForm, CustomAuthenticationForm, SearchForm, PerfilForm
from datetime import date, timedelta, datetime
from django.utils import timezone
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages

from django.template.loader import render_to_string
from weasyprint import HTML
from django.http import HttpResponse, JsonResponse
import pandas as pd
from django.core.paginator import Paginator
from django.db.models import Q
from django.contrib.auth.views import LoginView
from openpyxl.styles import Font, PatternFill
from .export_utils import export_documentos_pdf
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash


def is_admin(user):
    return user.is_superuser

def get_local_do_gestor(user):
    if user.is_superuser:
        return None  # superuser: acesso total
    gestor = GestorLocal.objects.filter(user=user).first()
    if gestor:
        # Retorna todos os locais do gestor (ManyToMany)
        return gestor.locais.all()
    return 'NO_PERMISSION'

@login_required
def no_permission(request):
    return render(request, 'no_permission.html')

@login_required
@check_permission
def dashboard(request):
    hoje = timezone.now().date()
    local_gestor = get_local_do_gestor(request.user)

    documentos = Documento.objects.all()
    if local_gestor and local_gestor != 'NO_PERMISSION':
        documentos = documentos.filter(funcionario__local_mobilizacao__in=local_gestor)

    funcionario_id = request.GET.get('funcionario')
    local_id = request.GET.get('local')
    tipo = request.GET.get('tipo')
    status = request.GET.get('filtro', 'todos')
    funcionario_status = request.GET.get('funcionario_status', 'ativos')  # novo filtro

    # Filtro de funcionários ativos/inativos/todos
    if funcionario_status == 'ativos':
        documentos = documentos.filter(funcionario__ativo=True)
    elif funcionario_status == 'inativos':
        documentos = documentos.filter(funcionario__ativo=False)
    # se for 'todos', não filtra

    if funcionario_id:
        documentos = documentos.filter(funcionario__id=funcionario_id)
    if local_id:
        documentos = documentos.filter(funcionario__local_mobilizacao__id=local_id)
    if tipo:
        documentos = documentos.filter(tipo_documento=tipo)

    if status == 'vencidos':
        documentos = documentos.filter(data_validade__lt=hoje)
    elif status == 'criticos':
        documentos = documentos.filter(data_validade__range=[hoje, hoje + timedelta(days=10)])
    elif status == 'vencendo':
        documentos = documentos.filter(data_validade__range=[hoje + timedelta(days=11), hoje + timedelta(days=40)])

    for d in documentos:
        d.dias_restantes = (d.data_validade - hoje).days

    # PAGINAÇÃO
    paginator = Paginator(documentos.order_by('-created_at'), 10)  # 10 por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Serializar documentos da página para o JS
    initial_documents = [
        {
            'id': doc.id,
            'funcionario_nome': doc.funcionario.nome,
            'nome_documento': doc.nome_documento,
            'tipo_documento': doc.tipo_documento,
            'data_emissao': doc.data_emissao.strftime('%d/%m/%Y'),
            'data_validade': doc.data_validade.strftime('%d/%m/%Y'),
            'dias_restantes': (doc.data_validade - hoje).days,
            'arquivo_url': doc.arquivo.url if doc.arquivo else None,
        }
        for doc in page_obj
    ]

    total_docs = documentos.count()
    docs_vencidos = documentos.filter(data_validade__lt=hoje).count()
    docs_criticos = documentos.filter(data_validade__range=[hoje, hoje + timedelta(days=10)]).count()
    docs_vencendo = documentos.filter(data_validade__range=[hoje + timedelta(days=11), hoje + timedelta(days=40)]).count()
    docs_em_dia = documentos.filter(data_validade__gt=hoje + timedelta(days=40)).count()

    # Dados para o modal de exportação
    funcionarios = Funcionario.objects.all()
    locais = LocalMobilizacao.objects.all()
    tipos = Documento.objects.values_list('tipo_documento', flat=True).distinct()

    # Documentos por status
    documentos_vencidos = Documento.objects.filter(data_validade__lt=hoje)
    documentos_criticos = Documento.objects.filter(
        data_validade__lte=hoje + timedelta(days=7),
        data_validade__gt=hoje
    )
    documentos_em_dia = Documento.objects.filter(data_validade__gt=hoje + timedelta(days=7))
    
    # Dados para o gráfico de tipos
    tipos_documentos = Documento.objects.values_list('tipo_documento', flat=True).distinct()
    quantidade_por_tipo = []
    for tipo in tipos_documentos:
        quantidade = Documento.objects.filter(tipo_documento=tipo).count()
        quantidade_por_tipo.append(quantidade)
    
    # Próximos 5 documentos a vencer
    proximos_vencer = Documento.objects.filter(data_validade__gte=hoje).order_by('data_validade')[:5]
    for doc in proximos_vencer:
        doc.dias_restantes = (doc.data_validade - hoje).days

    # KPIs globais
    docs_vencidos = documentos.filter(data_validade__lt=hoje).count()
    docs_criticos = documentos.filter(data_validade__range=[hoje, hoje + timedelta(days=10)]).count()
    docs_vencendo = documentos.filter(data_validade__range=[hoje + timedelta(days=11), hoje + timedelta(days=40)]).count()
    docs_em_dia = documentos.filter(data_validade__gt=hoje + timedelta(days=40)).count()
    total_docs = documentos.count()

    context = {
        'page_obj': page_obj,
        'total_documentos': total_docs,
        'docs_vencidos': docs_vencidos,
        'docs_criticos': docs_criticos,
        'docs_vencendo': docs_vencendo,
        'docs_em_dia': docs_em_dia,
        'funcionarios': funcionarios,
        'locais': locais,
        'tipos': tipos,
        'tipos_documentos': list(tipos_documentos),
        'quantidade_por_tipo': quantidade_por_tipo,
        'proximos_vencer': proximos_vencer,
        'filtro': status,
        'funcionario_status': funcionario_status,
        'alerta_vencidos': docs_vencidos,
        'alerta_criticos': docs_criticos,
        'alerta_vencendo': docs_vencendo,
        'initial_documents': initial_documents,
    }
    return render(request, 'dashboard.html', context)


@login_required
@check_permission
def funcionario_detail(request, funcionario_id):
    funcionario = get_object_or_404(Funcionario, id=funcionario_id)
    local_gestor = get_local_do_gestor(request.user)

    if local_gestor and local_gestor != 'NO_PERMISSION' and not funcionario.local_mobilizacao.filter(id__in=local_gestor).exists():
        return redirect('funcionario_list')

    documentos = Documento.objects.filter(funcionario=funcionario)
    hoje = date.today()
    for doc in documentos:
        doc.dias_restantes = (doc.data_validade - hoje).days

    return render(request, 'funcionario_detail.html', {
        'funcionario': funcionario,
        'documentos': documentos,
        'today': hoje,
    })

@login_required
@check_permission
def documentos_funcionario(request, funcionario_id):
    funcionario = get_object_or_404(Funcionario, id=funcionario_id)
    documentos = Documento.objects.filter(funcionario=funcionario)
    return render(request, 'documentos_funcionario.html', {'funcionario': funcionario, 'documentos': documentos})

@login_required
@check_permission
def funcionario_list(request):
    local_gestor = get_local_do_gestor(request.user)

    funcionarios = Funcionario.objects.all()
    if local_gestor and local_gestor != 'NO_PERMISSION':
        funcionarios = funcionarios.filter(local_mobilizacao__in=local_gestor)

    # Filtros
    nome = request.GET.get('nome', '')
    matricula = request.GET.get('matricula', '')
    status = request.GET.get('status', '')

    if nome:
        funcionarios = funcionarios.filter(nome__icontains=nome)
    if matricula:
        funcionarios = funcionarios.filter(matricula__icontains=matricula)
    if status == 'ativo':
        funcionarios = funcionarios.filter(ativo=True)
    elif status == 'inativo':
        funcionarios = funcionarios.filter(ativo=False)

    total_funcionarios = funcionarios.count()  

    
    # Paginação
    paginator = Paginator(funcionarios.order_by('id'), 10)  # 10 por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Dados para o modal de exportação
    locais = LocalMobilizacao.objects.all()
    if local_gestor and local_gestor != 'NO_PERMISSION':
        locais = locais.filter(id__in=[l.id for l in local_gestor])

    context = {
        'page_obj': page_obj,
        'nome': nome,
        'matricula': matricula,
        'locais': locais,
        'total_funcionarios': total_funcionarios,  # <--
    }
    return render(request, 'funcionario_list.html', context)

@login_required
@check_permission
def funcionario_create(request):
    form = FuncionarioForm(request.POST or None, user=request.user)
    if form.is_valid():
        funcionario = form.save(commit=False)
        funcionario.criado_por = request.user
        funcionario.atualizado_por = request.user
        form.save()
        form.save_m2m()  # Salva as relações ManyToMany
        messages.success(request, 'Funcionário criado com sucesso!')
        return redirect('funcionario_list')
    return render(request, 'funcionario_form.html', {'form': form})


@login_required
@check_permission
def funcionario_update(request, pk):
    funcionario = get_object_or_404(Funcionario, pk=pk)
    local_gestor = get_local_do_gestor(request.user)
    if local_gestor and local_gestor != 'NO_PERMISSION' and not funcionario.local_mobilizacao.filter(id__in=local_gestor).exists():
        return redirect('funcionario_list')

    form = FuncionarioForm(request.POST or None, instance=funcionario, user=request.user)
    if form.is_valid():
        funcionario = form.save(commit=False)
        funcionario.atualizado_por = request.user
        form.save()
        form.save_m2m()
        messages.success(request, 'Funcionário atualizado com sucesso!')
        return redirect('funcionario_list')
    return render(request, 'funcionario_form.html', {'form': form})



@login_required
@check_permission
def funcionario_delete(request, pk):
    funcionario = get_object_or_404(Funcionario, pk=pk)
    local_gestor = get_local_do_gestor(request.user)
    # Ajuste: local_gestor é um queryset de locais; usa id__in para checar todos os locais
    if local_gestor and local_gestor != 'NO_PERMISSION' and not funcionario.local_mobilizacao.filter(id__in=local_gestor).exists():
        return redirect('funcionario_list')

    if request.method == 'POST':
        funcionario.delete()
        messages.success(request, 'Funcionário excluído com sucesso!')
        return redirect('funcionario_list')
    return render(request, 'funcionario_confirm_delete.html', {'funcionario': funcionario})

@login_required
@check_permission
def documento_create(request, funcionario_id):
    funcionario = get_object_or_404(Funcionario, id=funcionario_id)
    local_gestor = get_local_do_gestor(request.user)
    if local_gestor and local_gestor != 'NO_PERMISSION' and not funcionario.local_mobilizacao.filter(id__in=local_gestor).exists():
        return redirect('funcionario_list')

    form = DocumentoForm(request.POST or None, request.FILES or None)
    if form.is_valid():
        documento = form.save(commit=False)
        documento.funcionario = funcionario
        documento.criado_por = request.user
        documento.atualizado_por = request.user
        documento.save()
        messages.success(request, 'Documento criado com sucesso!')
        return redirect('funcionario_detail', funcionario_id=funcionario.id)
    return render(request, 'documento_form.html', {'form': form, 'funcionario': funcionario})


@login_required
@check_permission
def documento_update(request, documento_id):
    documento = get_object_or_404(Documento, id=documento_id)
    funcionario = documento.funcionario
    local_gestor = get_local_do_gestor(request.user)
    if local_gestor and local_gestor != 'NO_PERMISSION' and not funcionario.local_mobilizacao.filter(id__in=local_gestor).exists():
        return redirect('funcionario_list')

    form = DocumentoForm(request.POST or None, request.FILES or None, instance=documento)
    if form.is_valid():
        documento = form.save(commit=False)
        documento.atualizado_por = request.user
        documento.save()
        messages.success(request, 'Documento atualizado com sucesso!')
        return redirect('funcionario_detail', funcionario_id=funcionario.id)
    return render(request, 'documento_form.html', {'form': form, 'funcionario': funcionario})


@login_required
@check_permission
def documento_delete(request, documento_id):
    documento = get_object_or_404(Documento, id=documento_id)
    funcionario = documento.funcionario
    local_gestor = get_local_do_gestor(request.user)
    if local_gestor and local_gestor != 'NO_PERMISSION' and not funcionario.local_mobilizacao.filter(id__in=local_gestor).exists():
        return redirect('funcionario_list')

    if request.method == 'POST':
        documento.delete()
        messages.success(request, 'Documento excluído com sucesso!')
        return redirect('funcionario_detail', funcionario_id=funcionario.id)
    return render(request, 'documento_confirm_delete.html', {'documento': documento})

@login_required
@user_passes_test(is_admin)
def local_list(request):
    busca = request.GET.get('busca', '')
    locais = LocalMobilizacao.objects.all()
    if busca:
        locais = locais.filter(nome__icontains=busca)
    paginator = Paginator(locais.order_by('nome'), 10)  # 10 por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'local_list.html', {'page_obj': page_obj, 'busca': busca})

@login_required
@user_passes_test(is_admin)
def local_create(request):
    form = LocalMobilizacaoForm(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect('local_list')
    return render(request, 'local_form.html', {'form': form})

@login_required
@user_passes_test(is_admin)
def local_update(request, pk):
    local = get_object_or_404(LocalMobilizacao, pk=pk)
    form = LocalMobilizacaoForm(request.POST or None, instance=local)
    if form.is_valid():
        form.save()
        return redirect('local_list')
    return render(request, 'local_form.html', {'form': form})

@login_required
@user_passes_test(is_admin)
def local_delete(request, pk):
    local = get_object_or_404(LocalMobilizacao, pk=pk)
    if request.method == 'POST':
        local.delete()
        return redirect('local_list')
    return render(request, 'local_confirm_delete.html', {'local': local})

@login_required
@user_passes_test(is_admin)
def gestor_list(request):
    busca = request.GET.get('busca', '')

    gestores = GestorLocal.objects.all()
    if busca:
        gestores = gestores.filter(
            user__username__icontains=busca
        ) | gestores.filter(
            # Ajuste: GestorLocal possui relação ManyToMany com LocalMobilizacao através de 'locais'
            locais__nome__icontains=busca
        )

    paginator = Paginator(gestores.order_by('id'), 10)  # 10 por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'busca': busca
    }
    return render(request, 'gestor_list.html', context)

@login_required
@user_passes_test(is_admin)
def gestor_create(request):
    form = GestorLocalForm(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect('gestor_list')
    return render(request, 'gestor_form.html', {'form': form})

@login_required
@user_passes_test(is_admin)
def gestor_update(request, pk):
    gestor = get_object_or_404(GestorLocal, pk=pk)
    form = GestorLocalForm(request.POST or None, instance=gestor)
    if form.is_valid():
        form.save()
        return redirect('gestor_list')
    return render(request, 'gestor_form.html', {'form': form})

@login_required
@user_passes_test(is_admin)
def gestor_delete(request, pk):
    gestor = get_object_or_404(GestorLocal, pk=pk)
    if request.method == 'POST':
        gestor.delete()
        return redirect('gestor_list')
    return render(request, 'gestor_confirm_delete.html', {'gestor': gestor})


@login_required
@check_permission
def logs_view(request):
    filtro = request.GET.get('filtro', 'funcionarios')  #  padrão funcionários

    funcionarios = Funcionario.objects.none()
    documentos = Documento.objects.none()

    if filtro == 'funcionarios':
        funcionarios = Funcionario.objects.all().order_by('-updated_at')
    elif filtro == 'documentos':
        documentos = Documento.objects.all().order_by('-updated_at')

    # Paginação
    paginator_func = Paginator(funcionarios, 20)
    page_func = paginator_func.get_page(request.GET.get('page_func'))

    paginator_doc = Paginator(documentos, 20)
    page_doc = paginator_doc.get_page(request.GET.get('page_doc'))

    context = {
        'filtro': filtro,
        'page_func': page_func,
        'page_doc': page_doc,
    }
    return render(request, 'logs.html', context)

@login_required
@check_permission
def export_funcionarios_pdf(request):
    funcionarios = Funcionario.objects.all().order_by('nome')
    html_string = render_to_string('funcionarios_pdf.html', {'funcionarios': funcionarios})
    html = HTML(string=html_string)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="funcionarios.pdf"'
    html.write_pdf(response)
    return response

@login_required
@check_permission
def export_documentos_pdf(request):
    documentos = Documento.objects.all().order_by('nome_documento')
    html_string = render_to_string('documentos_pdf.html', {'documentos': documentos})
    html = HTML(string=html_string)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="documentos.pdf"'
    html.write_pdf(response)
    return response


@login_required
@check_permission
def export_funcionarios_excel(request):
    # Obtém os filtros da URL
    status = request.GET.get('status')
    local_id = request.GET.get('local')
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    colunas = request.GET.getlist('colunas', ['nome', 'matricula', 'ativo', 'local', 'criado_por', 'criado_em'])

    # Filtra os funcionários
    funcionarios = Funcionario.objects.all()
    
    if status:
        funcionarios = funcionarios.filter(ativo=status == 'ativo')
    if local_id:
        funcionarios = funcionarios.filter(local_mobilizacao__id=local_id)
    if data_inicio:
        funcionarios = funcionarios.filter(created_at__gte=data_inicio)
    if data_fim:
        funcionarios = funcionarios.filter(created_at__lte=data_fim)

    # Prepara os dados para exportação
    data = []
    for func in funcionarios:
        func_data = {}
        if 'nome' in colunas:
            func_data['Nome'] = func.nome
        if 'matricula' in colunas:
            func_data['Matrícula'] = func.matricula
        if 'ativo' in colunas:
            func_data['Ativo'] = 'Sim' if func.ativo else 'Não'
        if 'local' in colunas:
            func_data['Local'] = ", ".join([l.nome for l in func.local_mobilizacao.all()])
        if 'criado_por' in colunas:
            func_data['Criado por'] = func.criado_por.username if func.criado_por else '---'
        if 'criado_em' in colunas:
            func_data['Criado em'] = func.created_at.strftime('%d/%m/%Y %H:%M')
        if 'atualizado_por' in colunas:
            func_data['Atualizado por'] = func.atualizado_por.username if func.atualizado_por else '---'
        if 'atualizado_em' in colunas:
            func_data['Atualizado em'] = func.updated_at.strftime('%d/%m/%Y %H:%M')
        data.append(func_data)

    df = pd.DataFrame(data)

    # Cria resposta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="funcionarios.xlsx"'

    # Escreve com cabeçalho e formatação
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        sheet_name = 'Funcionários'
        df.to_excel(writer, index=False, startrow=2, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
        
        # Adiciona cabeçalho
        ws['A1'] = 'Relatório de Funcionários'
        ws['A2'] = f'Exportado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        
        # Formata o cabeçalho
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'].font = Font(italic=True)
        
        # Formata as colunas
        for idx, col in enumerate(df.columns):
            cell = ws.cell(row=3, column=idx+1)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

    return response

@login_required
@check_permission
def export_documentos_excel(request):
    """Exporta documentos para Excel com filtros"""
    # Obter parâmetros do formulário
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    status = request.GET.get('status')
    tipo = request.GET.get('tipo_doc')
    local_id = request.GET.get('local')
    funcionario_id = request.GET.get('funcionario')
    colunas = request.GET.getlist('colunas')
    
    # Filtrar documentos
    documentos = Documento.objects.all()
    
    if data_inicio:
        documentos = documentos.filter(data_validade__gte=data_inicio)
    if data_fim:
        documentos = documentos.filter(data_validade__lte=data_fim)
    if status and status != 'todos':
        hoje = timezone.now().date()
        if status == 'vencidos':
            documentos = documentos.filter(data_validade__lt=hoje)
        elif status == 'criticos':
            documentos = documentos.filter(
                data_validade__lte=hoje + timedelta(days=10),
                data_validade__gt=hoje
            )
        elif status == 'vencendo':
            documentos = documentos.filter(
                data_validade__lte=hoje + timedelta(days=40),
                data_validade__gt=hoje + timedelta(days=10)
            )
        elif status == 'em_dia':
            documentos = documentos.filter(data_validade__gt=hoje + timedelta(days=40))
    if tipo:
        documentos = documentos.filter(tipo_documento=tipo)
    if local_id:
        documentos = documentos.filter(funcionario__local_mobilizacao__id=local_id)
    if funcionario_id:
        documentos = documentos.filter(funcionario__id=funcionario_id)
    
    # Verificar formato de exportação
    formato = request.GET.get('formato', 'excel')
    
    if formato == 'pdf':
        return export_documentos_pdf(documentos, colunas)
    
    # Preparar dados para Excel
    data = []
    headers = []

    # Mapear nomes das colunas
    coluna_names = {
        'nome': 'Nome',
        'funcionario': 'Funcionário',
        'tipo': 'Tipo',
        'data_emissao': 'Data de Emissão',
        'data_validade': 'Data de Validade',
        'status': 'Status',
        'local': 'Local'
    }

    # Adicionar cabeçalhos
    for col in colunas:
        if col in coluna_names:
            headers.append(coluna_names[col])

    data.append(headers)

    # Adicionar dados
    for doc in documentos:
        row = []
        for col in colunas:
            if col == 'nome':
                row.append(doc.nome_documento)
            elif col == 'funcionario':
                row.append(doc.funcionario.nome)
            elif col == 'tipo':
                row.append(doc.tipo_documento)
            elif col == 'data_emissao':
                row.append(doc.data_emissao.strftime('%d/%m/%Y'))
            elif col == 'data_validade':
                row.append(doc.data_validade.strftime('%d/%m/%Y'))
            elif col == 'status':
                dias_restantes = (doc.data_validade - timezone.now().date()).days
                if dias_restantes < 0:
                    status = 'Vencido'
                elif dias_restantes <= 10:
                    status = 'Crítico'
                elif dias_restantes <= 40:
                    status = 'Vencendo'
                else:
                    status = 'Em Dia'
                row.append(status)
            elif col == 'local':
                if doc.funcionario and doc.funcionario.local_mobilizacao.exists():
                    row.append(doc.funcionario.local_mobilizacao.first().nome)
                else:
                    row.append('')
        data.append(row)

    # Criar DataFrame e exportar para Excel
    df = pd.DataFrame(data[1:], columns=data[0])
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = f'attachment; filename="documentos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    df.to_excel(response, index=False)

    return response

class CustomLoginView(LoginView):
    template_name = 'login.html'
    authentication_form = CustomAuthenticationForm

@login_required
@check_permission
def advanced_search(request):
    search_type = request.GET.get('search_type', 'documentos')
    resultados = []

    # None = superuser (acesso total); 'NO_PERMISSION' = sem vínculo
    local_gestor = get_local_do_gestor(request.user)

    if search_type == 'documentos':
        documentos = Documento.objects.all()
        if local_gestor and local_gestor != 'NO_PERMISSION':
            documentos = documentos.filter(
                funcionario__local_mobilizacao__in=local_gestor
            ).distinct()

        # Filtros
        nome = request.GET.get('nome')
        status = request.GET.get('status')
        data_inicio = request.GET.get('data_inicio')
        data_fim = request.GET.get('data_fim')

        if nome:
            documentos = documentos.filter(nome_documento__icontains=nome)

        if status:
            hoje = timezone.now().date()
            if status == 'vencido':
                documentos = documentos.filter(data_validade__lt=hoje)
            elif status == 'critico':
                documentos = documentos.filter(
                    data_validade__gte=hoje,
                    data_validade__lte=hoje + timedelta(days=10)
                )
            elif status == 'vencendo':
                documentos = documentos.filter(
                    data_validade__gt=hoje + timedelta(days=10),
                    data_validade__lte=hoje + timedelta(days=40)
                )
            elif status == 'em_dia':
                documentos = documentos.filter(data_validade__gt=hoje + timedelta(days=40))

        if data_inicio:
            documentos = documentos.filter(data_validade__gte=data_inicio)
        if data_fim:
            documentos = documentos.filter(data_validade__lte=data_fim)

        # dias_restantes
        hoje = timezone.now().date()
        for doc in documentos:
            doc.dias_restantes = (doc.data_validade - hoje).days

        resultados = documentos

    elif search_type == 'funcionarios':
        funcionarios = Funcionario.objects.all()
        if local_gestor and local_gestor != 'NO_PERMISSION':
            funcionarios = funcionarios.filter(
                local_mobilizacao__in=local_gestor
            ).distinct()

        # Filtros
        nome = request.GET.get('nome_funcionario')
        cargo = request.GET.get('cargo')
        status = request.GET.get('status_funcionario')

        if nome:
            funcionarios = funcionarios.filter(nome__icontains=nome)
        if cargo:
            funcionarios = funcionarios.filter(cargo__icontains=cargo)
        if status:
            if status == 'ativo':
                funcionarios = funcionarios.filter(ativo=True)
            elif status == 'inativo':
                funcionarios = funcionarios.filter(ativo=False)

        resultados = funcionarios

    elif search_type == 'mobilizacoes':
        mobilizacoes = LocalMobilizacao.objects.all()
        if local_gestor and local_gestor != 'NO_PERMISSION':
            # restringe às mobilizações (locais) que o gestor gerencia
            mobilizacoes = mobilizacoes.filter(
                id__in=[l.id for l in local_gestor]
            )

        # Filtros
        local = request.GET.get('local')
        data_inicio = request.GET.get('data_inicio_mob')
        data_fim = request.GET.get('data_fim_mob')
        status = request.GET.get('status_mobilizacao')

        if local:
            mobilizacoes = mobilizacoes.filter(nome__icontains=local)
        if data_inicio:
            mobilizacoes = mobilizacoes.filter(data_inicio__gte=data_inicio)
        if data_fim:
            mobilizacoes = mobilizacoes.filter(data_fim__lte=data_fim)
        if status:
            mobilizacoes = mobilizacoes.filter(status=status)

        resultados = mobilizacoes

    # Paginação (mantida)
    paginator = Paginator(resultados, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'search_type': search_type,
        'resultados': page_obj,
        'form': SearchForm(),
        'page_obj': page_obj
    }
    return render(request, 'advanced_search.html', context)

@login_required
def get_filtered_dashboard_data(request):
    status = request.GET.get('status', 'Todos')
    funcionario_status = request.GET.get('funcionario_status', 'ativos')
    hoje = timezone.now().date()
    local_gestor = get_local_do_gestor(request.user)

    documentos = Documento.objects.all()
    if local_gestor and local_gestor != 'NO_PERMISSION':
        documentos = documentos.filter(funcionario__local_mobilizacao__in=local_gestor)

    # Filtro de funcionários ativos/inativos/todos
    if funcionario_status == 'ativos':
        documentos = documentos.filter(funcionario__ativo=True)
    elif funcionario_status == 'inativos':
        documentos = documentos.filter(funcionario__ativo=False)
    # se for 'todos', não filtra

    if status == 'Vencidos':
        documentos = documentos.filter(data_validade__lt=hoje)
    elif status == 'Críticos':
        documentos = documentos.filter(data_validade__range=[hoje, hoje + timedelta(days=10)])
    elif status == 'Vencendo':
        documentos = documentos.filter(data_validade__range=[hoje + timedelta(days=11), hoje + timedelta(days=40)])
    elif status == 'Em Dia':
        documentos = documentos.filter(data_validade__gt=hoje + timedelta(days=40))

    # Dados para o gráfico de tipos
    tipos_documentos = documentos.values_list('tipo_documento', flat=True).distinct()
    quantidade_por_tipo = []
    for tipo in tipos_documentos:
        quantidade = documentos.filter(tipo_documento=tipo).count()
        quantidade_por_tipo.append(quantidade)

    # Dados para a tabela de documentos
    documentos_data = []
    for doc in documentos.order_by('-created_at'): # Ordenar para consistência com a paginação
        dias_restantes = (doc.data_validade - hoje).days
        documentos_data.append({
            'id': doc.id,
            'funcionario_nome': doc.funcionario.nome,
            'nome_documento': doc.nome_documento,
            'tipo_documento': doc.tipo_documento,
            'data_emissao': doc.data_emissao.strftime('%d/%m/%Y'),
            'data_validade': doc.data_validade.strftime('%d/%m/%Y'),
            'dias_restantes': dias_restantes,
            'arquivo_url': doc.arquivo.url if doc.arquivo else None,
        })

    # Dados para os KPIs
    docs_vencidos = documentos.filter(data_validade__lt=hoje).count()
    docs_criticos = documentos.filter(data_validade__range=[hoje, hoje + timedelta(days=10)]).count()
    docs_vencendo = documentos.filter(data_validade__range=[hoje + timedelta(days=11), hoje + timedelta(days=40)]).count()
    docs_em_dia = documentos.filter(data_validade__gt=hoje + timedelta(days=40)).count()
    total_documentos = documentos.count()

    # Dados para gráfico de documentos por local
    locais_labels = []
    locais_data = []
    locais = LocalMobilizacao.objects.all()
    for local in locais:
        count = documentos.filter(funcionario__local_mobilizacao=local).count()
        locais_labels.append(local.nome)
        locais_data.append(count)

    return JsonResponse({
        'tipos_documentos': list(tipos_documentos),
        'quantidade_por_tipo': quantidade_por_tipo,
        'documentos': documentos_data,
        'status_data': [docs_vencidos, docs_criticos, docs_vencendo, docs_em_dia],
        'kpis': {
            'vencidos': docs_vencidos,
            'criticos': docs_criticos,
            'vencendo': docs_vencendo,
            'em_dia': docs_em_dia,
            'total': total_documentos
        },
        'locais_labels': locais_labels,
        'locais_data': locais_data,
    })

@login_required
@check_permission
def alertas(request):
    hoje = timezone.now().date()
    local_gestor = get_local_do_gestor(request.user)
    documentos = Documento.objects.all()
    if local_gestor and local_gestor != 'NO_PERMISSION':
        documentos = documentos.filter(funcionario__local_mobilizacao__in=local_gestor)
    group = request.GET.get('group', 'local')
    total_vencidos = documentos.filter(data_validade__lt=hoje).count()
    total_criticos = documentos.filter(data_validade__range=[hoje, hoje + timedelta(days=10)]).count()
    total_vencendo = documentos.filter(data_validade__range=[hoje + timedelta(days=11), hoje + timedelta(days=40)]).count()
    agrupados = {}
    if group == 'funcionario':
        for func in documentos.values_list('funcionario', flat=True).distinct():
            funcionario = Funcionario.objects.get(id=func)
            docs_func = documentos.filter(funcionario=funcionario)
            vencidos = docs_func.filter(data_validade__lt=hoje)
            criticos = docs_func.filter(data_validade__range=[hoje, hoje + timedelta(days=10)])
            vencendo = docs_func.filter(data_validade__range=[hoje + timedelta(days=11), hoje + timedelta(days=40)])
            if vencidos.exists() or criticos.exists() or vencendo.exists():
                agrupados[funcionario.nome] = {
                    'vencidos': vencidos,
                    'criticos': criticos,
                    'vencendo': vencendo
                }
    else:
        for local in LocalMobilizacao.objects.all():
            docs_local = documentos.filter(funcionario__local_mobilizacao=local)
            vencidos = docs_local.filter(data_validade__lt=hoje)
            criticos = docs_local.filter(data_validade__range=[hoje, hoje + timedelta(days=10)])
            vencendo = docs_local.filter(data_validade__range=[hoje + timedelta(days=11), hoje + timedelta(days=40)])
            if vencidos.exists() or criticos.exists() or vencendo.exists():
                agrupados[local.nome] = {
                    'vencidos': vencidos,
                    'criticos': criticos,
                    'vencendo': vencendo
                }
    context = {
        'agrupados': agrupados,
        'group': group,
        'total_vencidos': total_vencidos,
        'total_criticos': total_criticos,
        'total_vencendo': total_vencendo,
    }
    return render(request, 'alertas.html', context)

@login_required
def perfil_usuario(request):
    user = request.user
    perfil, _ = Perfil.objects.get_or_create(user=user)
    if request.method == 'POST':
        perfil_form = PerfilForm(request.POST, request.FILES, instance=perfil)
        senha_form = PasswordChangeForm(user, request.POST)
        if 'salvar_perfil' in request.POST and perfil_form.is_valid():
            perfil_form.save()
            messages.success(request, 'Perfil atualizado com sucesso!')
            return redirect('perfil_usuario')
        elif 'trocar_senha' in request.POST and senha_form.is_valid():
            user = senha_form.save()
            update_session_auth_hash(request, user)
            messages.success(request, 'Senha alterada com sucesso!')
            return redirect('perfil_usuario')
    else:
        perfil_form = PerfilForm(instance=perfil)
        senha_form = PasswordChangeForm(user)
    return render(request, 'perfil_usuario.html', {
        'perfil_form': perfil_form,
        'senha_form': senha_form,
        'perfil': perfil
    })
